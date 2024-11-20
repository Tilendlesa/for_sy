import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json

VERSION = 'v1.1.0'

# ============== 配置文件读取 ==============

# 需要忽视的同学序号
IGNORE_STUDENTS = []

# 五大项题目数量
LISTEN_NUM = 0
READING_NUM = 0
FILL_BLANK_NUM = 0
GRAMMER_NUM = 0
SPELL_NUM = 0

# ============== 其他配置文件依赖项 ==============
NEW_COLUMN_RANGE = []
CHOICE_QUESTION_NUM = 0

# ============== 其他全局变量 ==============

# 五大项题目名称
LISTEN_NAME = '听力'
READING_NAME = '阅读'
FILL_BLANK_NAME = '完形'
GRAMMER_NAME = '语法填空'
SPELL_NAME = '单词拼写'
CHOICE_QUESTION_NAME = '客观题'

new_column_num = 5 # 新增五列
new_column_name = [LISTEN_NAME, READING_NAME, FILL_BLANK_NAME, GRAMMER_NAME, SPELL_NAME]
new_column_need_to_mark = [25.5, 40, 0, 10.5, 0] # 低于这个分数标红


# 需要图红色的作文分数
WRITING_LOWEST_SCORE = 8

# 需要图红色的校级进退步
SCHOOL_FALL_BEGIND = -100

# 文件路径
input_file_path = 'input1.xlsx'
output_file_path_without_color = 'output_without_color.xlsx'
output_file_path_with_color = 'output_with_color.xlsx'
conf_path = 'conf.json'

# 定义颜色背景填充
color_fills = {
    'top1': PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"),   # 粉色
    '0-10': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),   # 浅蓝色
    '10-20': PatternFill(start_color="008000", end_color="008000", fill_type="solid"),  # 绿色
    '20-30': PatternFill(start_color="ADFF2F", end_color="ADFF2F", fill_type="solid"),  # 浅绿色
    '30-40': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # 黄色
    '40-50': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # 橙色
    '50-100': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), # 红色
}

def read_conf():
    global IGNORE_STUDENTS
    global LISTEN_NUM
    global READING_NUM
    global FILL_BLANK_NUM
    global GRAMMER_NUM
    global SPELL_NUM
    with open(os.path.join(get_base_path(), conf_path), 'r', encoding='utf-8') as file:
        data = json.load(file)
    
        IGNORE_STUDENTS = data['ignore_students']
        LISTEN_NUM = data['listening_num']
        READING_NUM = data['reading_num']
        FILL_BLANK_NUM = data['fill_blank_num']
        GRAMMER_NUM = data['grammer_num']
        SPELL_NUM = data['spell_num']

    global NEW_COLUMN_RANGE
    NEW_COLUMN_RANGE = [LISTEN_NUM, READING_NUM, FILL_BLANK_NUM, GRAMMER_NUM, SPELL_NUM]

    global CHOICE_QUESTION_NUM
    CHOICE_QUESTION_NUM = LISTEN_NUM + READING_NUM + FILL_BLANK_NUM

def get_base_path():
    if getattr(sys, 'frozen', False):
        # 如果是打包后的可执行文件
        base_path = os.path.dirname(sys.executable)
    else:
        # 如果是正常的脚本运行
        base_path = os.path.dirname(os.path.abspath(__file__))
    return base_path

def calc_score():
    # 读取 Excel 文件
    df = pd.read_excel(os.path.join(get_base_path(), input_file_path), sheet_name='得分明细', skiprows=1)

    # 删除指定的列
    columns_to_drop = ['准考证号', '自定义考号', '班级']
    df = df.drop(columns=columns_to_drop)

    # 合并进退步到得分明细表

    # 选择需要的列
    df_collect = pd.read_excel(os.path.join(get_base_path(), input_file_path), sheet_name='班级英语成绩汇总', skiprows=1)
    columns_to_add = df_collect[['序号', '校次进退步', '班次进退步']]

    # 合并数据表
    # 通过 '序号' 列进行合并，并保留第一张表的所有数据
    df = pd.merge(df, columns_to_add, on='序号', how='left')

    # 获取 '班次' 列的索引
    class_rank_index = df.columns.get_loc('班次')

    # 重新排列列的顺序
    columns = list(df.columns)
    # 将新列插入到 '班次' 列后面
    new_columns_order = columns[:class_rank_index + 1] + ['校次进退步', '班次进退步'] + columns[class_rank_index + 1:-2]
    df = df[new_columns_order]
    print('插入校次进退步和班次进退步完毕')


    # 排除校次为 '-' 的行，不计算
    df_valid = df[df['总分'] != '未扫，不计排名'].copy()

    # 确定插入新列的位置
    cols = df_valid.columns.tolist()
    insert_pos = cols.index('主观分') + 1

    # 读取需要忽略的学生
    # with open(os.path.join(get_base_path(), ignore_student_path), 'r') as file:
    #     # 读取每一行并移除换行符
    #     lines = [line.strip() for line in file]

    # # 将读取的内容转换为整数
    # ignore_students = [int(line) for line in lines]
    # mask = ~df_valid.index.isin(ignore_students)

    # 计算新列值
    prefix = 1
    for i in range(new_column_num):
        df_valid[new_column_name[i]] = 0
        for j in range(NEW_COLUMN_RANGE[i]):
            column_name = f'{prefix}'
            # df_valid.loc[mask, new_column_name[i]] += df_valid.loc[mask, column_name]
            df_valid[new_column_name[i]] += df_valid[column_name]
            prefix += 1
        # print('df[new_column_name[i]]: ' + str(df[new_column_name[i]]))
        cols.insert(insert_pos, new_column_name[i])
        df_valid = df_valid[cols]
        insert_pos = cols.index(new_column_name[i]) + 1
        print(new_column_name[i] + '分数计算完毕')


    # 将计算后的听力和阅读信息回填到原 DataFrame 中
    cols = df.columns.tolist()
    truly_pos = cols.index('主观分') + 1
    
    for i in range(new_column_num):
        df.insert(truly_pos, new_column_name[i], pd.NA)
        df.loc[df_valid.index, new_column_name[i]] = df_valid[new_column_name[i]]
        cols.insert(truly_pos, new_column_name[i])
        truly_pos = cols.index(new_column_name[i]) + 1

    for i in range(new_column_num):
        # 年级得分率%
        df.loc[df.index[-1], new_column_name[i]] = df.loc[df.index[-1], new_column_name[i]] / NEW_COLUMN_RANGE[i]
        # 班级得分率%
        df.loc[df.index[-3], new_column_name[i]] = df.loc[df.index[-3], new_column_name[i]] / NEW_COLUMN_RANGE[i]
    
    # 客观题为听力+阅读+完形/3
    cols = df.columns.tolist()
    insert_pos = cols.index('主观分') + 1
    cols.insert(insert_pos, CHOICE_QUESTION_NAME)
    df[CHOICE_QUESTION_NAME] = (df[LISTEN_NAME] + df[READING_NAME] + df[FILL_BLANK_NAME])
    df = df[cols]

    # 年级得分率%
    df.loc[df.index[-1], CHOICE_QUESTION_NAME] = df.loc[df.index[-1], CHOICE_QUESTION_NAME] / 3
    # 班级得分率%
    df.loc[df.index[-3], CHOICE_QUESTION_NAME] = df.loc[df.index[-3], CHOICE_QUESTION_NAME] / 3

    # 保存处理后的数据到一个新的 Excel 文件
    df.to_excel(os.path.join(get_base_path(), output_file_path_without_color), index=False)
    print(f"计算分数,处理后的文件已保存到: {os.path.join(get_base_path(), output_file_path_without_color)}")

def render_color():

    # 加载Excel文件
    wb = load_workbook(os.path.join(get_base_path(), output_file_path_without_color))
    ws = wb.active

    # # 平均分行
    # avg_score_row_index = ws.max_row - 3

    # 找到“总分”列的列号
    total_score_column = None
    for cell in ws[1]:  # 假设标题在第一行
        if cell.value == "总分":
            total_score_column = cell.column_letter
            break

    if total_score_column is None:
        raise ValueError("未找到'总分'列，请检查标题名称")


    for i in range(new_column_num):

        if new_column_need_to_mark[i] == 0:
            continue

        render_red_color_when_item_is_lower_than_score(ws, total_score_column, new_column_name[i], new_column_need_to_mark[i])

        # column = None
        # for cell in ws[1]:  # 假设标题在第一行
        #     if cell.value == new_column_name[i]:
        #         column = cell.column_letter
        #         break

        # if column is None:
        #     raise ValueError("未找到" + new_column_name[i] + "列，请检查标题名称")

        # 获取倒数第四行的班级均分
        # avg_score_cell = ws[f'{column}{avg_score_row_index}']
        # avg_score = avg_score_cell.value

        # # 获取分数并存入列表中
        # scores = []
        # for row in range(2, avg_score_row_index):
        #     if ws[f'{total_score_column}{row}'].value == '未扫，不计排名':
        #         continue
        #     score_cell = ws[f'{column}{row}']
        #     score = score_cell.value
        #     if score is not None:
        #         scores.append(score)

        # # 排序分数
        # scores_sorted = sorted(scores, reverse=True)

        # total_students = len(scores_sorted)
        # percentile_10_idx = total_students // 10
        # percentile_20_idx = total_students // 5
        # percentile_30_idx = 3 * total_students // 10
        # percentile_40_idx = 2 * total_students // 5

        # score_top1 = scores_sorted[0]
        # score_10 = scores_sorted[percentile_10_idx] if percentile_10_idx < total_students else 0
        # score_20 = scores_sorted[percentile_20_idx] if percentile_20_idx < total_students else 0
        # score_30 = scores_sorted[percentile_30_idx] if percentile_30_idx < total_students else 0
        # score_40 = scores_sorted[percentile_40_idx] if percentile_40_idx < total_students else 0

        # for row in range(2, avg_score_row_index):  # 从第2行到最后一行
        #     if ws[f'{total_score_column}{row}'].value == '未扫，不计排名':
        #         continue

        #     score_cell = ws[f'{column}{row}']
        #     score = score_cell.value

        #     if score < new_column_need_to_mark[i]:
        #         score_cell.fill = color_fills['50-100']
            # if score is not None:
            #     if score >= score_top1:
            #         score_cell.fill = color_fills['top1']
            #     elif score >= score_10:
            #         score_cell.fill = color_fills['0-10']
            #     elif score >= score_20:
            #         score_cell.fill = color_fills['10-20']
            #     elif score >= score_30:
            #         score_cell.fill = color_fills['20-30']
            #     elif score >= score_40:
            #         score_cell.fill = color_fills['30-40']
            #     elif score >= avg_score:
            #         score_cell.fill = color_fills['40-50']
            #     else:
            #         score_cell.fill = color_fills['50-100']

        # print(new_column_name[i] + "涂色完毕")

    # 英语作文涂色
    render_red_color_when_item_is_lower_than_score(ws, total_score_column, '英语作文', WRITING_LOWEST_SCORE)

    # 校次进退步涂色
    render_red_color_when_item_is_lower_than_score(ws, total_score_column, '校次进退步', SCHOOL_FALL_BEGIND)

    # 保存修改后的Excel文件
    wb.save(os.path.join(get_base_path(), output_file_path_with_color))
    print(f"添加颜色,处理后的文件已保存到: {os.path.join(get_base_path(), output_file_path_with_color)}")

def render_red_color_when_item_is_lower_than_score(ws, total_score_column, item_name, lowest_score):

    column = None
    for cell in ws[1]:  # 假设标题在第一行
        if cell.value == item_name:
            column = cell.column_letter
            break

    if column is None:
        raise ValueError("未找到" + "校次进退步列，请检查标题名称")

     # 平均分行
    avg_score_row_index = ws.max_row - 3
    for row in range(2, avg_score_row_index):  # 从第2行到最后一行
        if ws[f'{total_score_column}{row}'].value == '未扫，不计排名':
            continue
        score_cell = ws[f'{column}{row}']
        score = score_cell.value
        if score < lowest_score:
            score_cell.fill = color_fills['50-100']

    print(item_name + "涂色完毕")

def love():
    read_conf()
    calc_score()
    render_color()

if __name__ == "__main__":
    print("版本号:" + VERSION)
    love()
 

 
